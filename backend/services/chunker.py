"""File chunking service.

This module contains helpers for splitting uploaded documents into
smaller chunks suitable for processing by the AI model. PDF files are
split into 100‑page blocks, DOCX files are treated as single chunks
and each sheet of an XLSX file becomes its own chunk. For PDFs,
image‑only pages are detected and marked accordingly.

The functions defined here expect an active SQLAlchemy async session
and will create `Chunk` records linked to the provided document and
session IDs. They also update the document's `total_pages` and
`total_chunks` fields based on the number of chunks generated.
"""

from __future__ import annotations

import os
from typing import Optional

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

try:
    import fitz  # type: ignore
except Exception:  # pragma: no cover
    fitz = None

try:
    import docx  # type: ignore
except Exception:  # pragma: no cover
    docx = None

try:
    import openpyxl  # type: ignore
except Exception:  # pragma: no cover
    openpyxl = None

from ..models import Document, Chunk


async def chunk_pdf(file_path: str, document: Document, session: AsyncSession) -> None:
    """Split a PDF into 100‑page chunks and create Chunk records.

    Parameters
    ----------
    file_path: str
        Path to the saved PDF file on disk.
    document: Document
        Document ORM instance to which the chunks belong.
    session: AsyncSession
        Active database session for persisting Chunk objects.
    """
    if fitz is None:
        raise RuntimeError("PyMuPDF is not available in this environment.")
    pdf = fitz.open(file_path)
    total_pages = pdf.page_count
    chunk_index = 0
    for start in range(0, total_pages, 100):
        end = min(start + 100, total_pages)
        text = ""
        for i in range(start, end):
            page_num = i + 1
            page = pdf.load_page(i)
            page_text = (page.get_text() or "").strip()
            text += f"\n\n--- PAGE {page_num} ---\n\n"
            if len(page_text) < 50:
                text += f"[IMAGE-ONLY PAGE {page_num} — text extraction unavailable]"
            else:
                text += page_text
        token_estimate = len(text) // 4 if text else 0
        chunk = Chunk(
            document_id=document.id,
            session_id=document.session_id,
            chunk_index=chunk_index,
            page_start=start + 1,
            page_end=end,
            text_content=text,
            token_estimate=token_estimate,
        )
        session.add(chunk)
        chunk_index += 1
    document.total_pages = total_pages
    document.total_chunks = chunk_index


async def chunk_docx(file_path: str, document: Document, session: AsyncSession) -> None:
    """Create a single chunk from a DOCX document.

    Parameters
    ----------
    file_path: str
        Path to the saved DOCX file.
    document: Document
        Document ORM instance.
    session: AsyncSession
        Database session.
    """
    if docx is None:
        raise RuntimeError("python-docx is not available in this environment.")
    doc = docx.Document(file_path)
    paragraphs = [p.text for p in doc.paragraphs if p.text]
    text = "\n\n".join(paragraphs)
    token_estimate = len(text) // 4 if text else 0
    chunk = Chunk(
        document_id=document.id,
        session_id=document.session_id,
        chunk_index=0,
        page_start=1,
        page_end=None,
        text_content=text,
        token_estimate=token_estimate,
    )
    session.add(chunk)
    document.total_pages = None
    document.total_chunks = 1


async def chunk_xlsx(file_path: str, document: Document, session: AsyncSession) -> None:
    """Create one chunk per sheet from an XLSX workbook.

    Parameters
    ----------
    file_path: str
        Path to the XLSX file.
    document: Document
        Document ORM instance.
    session: AsyncSession
        Database session.
    """
    if openpyxl is None:
        raise RuntimeError("openpyxl is not available in this environment.")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    total_chunks = 0
    for idx, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        lines: list[str] = []
        for row in ws.iter_rows(values_only=True):
            row_text = "\t".join([str(cell) if cell is not None else "" for cell in row]).rstrip()
            lines.append(row_text)
        sheet_text = f"Sheet: {sheet_name}\n" + "\n".join(lines)
        token_estimate = len(sheet_text) // 4 if sheet_text else 0
        chunk = Chunk(
            document_id=document.id,
            session_id=document.session_id,
            chunk_index=idx,
            page_start=None,
            page_end=None,
            text_content=sheet_text,
            token_estimate=token_estimate,
        )
        session.add(chunk)
        total_chunks += 1
    document.total_pages = None
    document.total_chunks = total_chunks


async def process_document_file(
    document: Document,
    file_path: str,
    session: AsyncSession,
) -> None:
    """Dispatch to the appropriate chunker based on document format."""
    fmt = (document.format or "").lower()
    if fmt == 'pdf':
        await chunk_pdf(file_path, document, session)
    elif fmt == 'docx':
        await chunk_docx(file_path, document, session)
    elif fmt == 'xlsx':
        await chunk_xlsx(file_path, document, session)
    else:
        raise ValueError(f"Unsupported document format: {document.format}")