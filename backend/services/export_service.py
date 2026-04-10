"""Export service for findings and clarifications.

This module contains helper functions to assemble an Excel workbook
containing the current session findings and the full clarification log.
The resulting workbook includes two worksheets: the first sheet
contains findings filtered by the user‑provided query parameters and
the second sheet lists every clarification question and its status.

The export is built using ``openpyxl`` and returned as a bytes
object that can be written directly to a ``StreamingResponse``.  The
calling code is responsible for setting appropriate HTTP headers.
"""

from __future__ import annotations

import io
from typing import Any, Dict, List, Optional

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from ..models import Finding as FindingModel
from ..models import Chunk as ChunkModel
from ..models import Document as DocumentModel
from ..models import Clarification as ClarificationModel


async def _load_chunk_and_doc_maps(
    findings: List[FindingModel], db: AsyncSession
) -> tuple[Dict[Any, ChunkModel], Dict[Any, str]]:
    """Preload chunks and documents to avoid N+1 queries.

    Returns a tuple (chunk_map, doc_map) where chunk_map maps chunk IDs
    to Chunk objects and doc_map maps document IDs to original file names.
    """
    chunk_ids = {f.chunk_id for f in findings if f.chunk_id is not None}
    chunk_map: Dict[Any, ChunkModel] = {}
    doc_map: Dict[Any, str] = {}
    if chunk_ids:
        res_chunks = await db.execute(select(ChunkModel).where(ChunkModel.id.in_(chunk_ids)))
        for ch in res_chunks.scalars().all():
            chunk_map[ch.id] = ch
        doc_ids = {ch.document_id for ch in chunk_map.values() if ch.document_id is not None}
        if doc_ids:
            res_docs = await db.execute(
                select(DocumentModel.id, DocumentModel.original_filename).where(DocumentModel.id.in_(doc_ids))
            )
            for doc_id, filename in res_docs:
                doc_map[doc_id] = filename
    return chunk_map, doc_map


async def export_findings_to_excel(
    session_id: str,
    filters: Optional[Dict[str, Any]],
    db: AsyncSession,
) -> bytes:
    """Build an Excel workbook containing filtered findings and all clarifications.

    Parameters
    ----------
    session_id: str
        Identifier of the session whose findings should be exported.
    filters: dict or None
        A dictionary of optional filters.  Supported keys are
        ``severity``, ``category``, ``document_id``, ``confidence`` and
        ``confirmed``.  Any filter that is ``None`` is ignored.
    db: AsyncSession
        Database session used to fetch records.

    Returns
    -------
    bytes
        The binary content of the generated Excel workbook.
    """
    # Build the base query for findings within the given session
    stmt = select(FindingModel).where(FindingModel.session_id == session_id)
    if filters:
        severity = filters.get("severity")
        category = filters.get("category")
        document_id = filters.get("document_id")
        confidence = filters.get("confidence")
        confirmed = filters.get("confirmed")
        if severity:
            stmt = stmt.where(FindingModel.severity == severity)
        if category:
            stmt = stmt.where(FindingModel.category == category)
        if document_id:
            # Filter by the finding.document_id field when present
            stmt = stmt.where(FindingModel.document_id == document_id)
        if confidence:
            stmt = stmt.where(FindingModel.confidence == confidence)
        if confirmed is not None:
            ci = str(confirmed).strip().lower()
            if ci == 'true':
                stmt = stmt.where(FindingModel.confirmed_correct.is_(True))
            elif ci == 'false':
                stmt = stmt.where(FindingModel.confirmed_correct.is_(False))
            elif ci in ('none', 'null', ''):
                stmt = stmt.where(FindingModel.confirmed_correct.is_(None))
            else:
                # Ignore invalid values silently for export
                pass
    result = await db.execute(stmt)
    findings: List[FindingModel] = result.scalars().all()

    # Preload chunk and document information for findings
    chunk_map, doc_map = await _load_chunk_and_doc_maps(findings, db)

    # Build workbook
    wb = Workbook()
    ws_findings = wb.active
    ws_findings.title = "Findings"
    # Header row for findings sheet
    header1 = [
        "Finding ID",
        "Finding Label",
        "Document",
        "Page Range/Section",
        "Original Text",
        "Category",
        "Comment",
        "Recommendation",
        "Severity",
        "Source Reference",
        "Confidence",
        "Confirmed",
        "Confirmed At",
    ]
    ws_findings.append(header1)
    # Populate rows
    for f in findings:
        doc_name = None
        page_range = None
        if f.chunk_id:
            ch = chunk_map.get(f.chunk_id)
            if ch:
                doc_name = doc_map.get(ch.document_id)
                if ch.page_start is not None and ch.page_end is not None:
                    page_range = f"{ch.page_start}-{ch.page_end}"
                elif ch.page_start is not None:
                    page_range = str(ch.page_start)
        elif f.document_id:
            doc_name = doc_map.get(f.document_id)
        ws_findings.append(
            [
                str(f.id),
                f.finding_label,
                doc_name or "",
                f.page_section_table or page_range or "",
                f.original_text or "",
                f.category or "",
                f.comment or "",
                f.recommendation or "",
                f.severity or "",
                f.source_reference or "",
                f.confidence or "",
                ("Yes" if f.confirmed_correct else "No")
                if f.confirmed_correct is not None
                else "",
                f.confirmed_at.isoformat() if f.confirmed_at else "",
            ]
        )

    # Auto-adjust column widths for findings sheet
    for i, column_title in enumerate(header1, start=1):
        max_length = len(column_title)
        for cell in ws_findings[get_column_letter(i)]:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws_findings.column_dimensions[get_column_letter(i)].width = max_length + 2

    # Clarifications sheet
    ws_clar = wb.create_sheet(title="Clarifications")
    header2 = [
        "Clarification ID",
        "Document",
        "Page Range",
        "Question",
        "Answer",
        "Status",
        "Created At",
        "Answered At",
    ]
    ws_clar.append(header2)
    # Query all clarifications for the session
    res_clar = await db.execute(
        select(ClarificationModel).where(ClarificationModel.session_id == session_id)
    )
    clarifications: List[ClarificationModel] = res_clar.scalars().all()
    if clarifications:
        # Preload chunk and document information for clarifications
        clar_chunk_ids = {cl.chunk_id for cl in clarifications if cl.chunk_id is not None}
        clar_chunk_map: Dict[Any, ChunkModel] = {}
        clar_doc_map: Dict[Any, str] = {}
        if clar_chunk_ids:
            res_chunks = await db.execute(select(ChunkModel).where(ChunkModel.id.in_(clar_chunk_ids)))
            for ch in res_chunks.scalars().all():
                clar_chunk_map[ch.id] = ch
            clar_doc_ids = {ch.document_id for ch in clar_chunk_map.values() if ch.document_id is not None}
            if clar_doc_ids:
                res_docs = await db.execute(
                    select(DocumentModel.id, DocumentModel.original_filename).where(DocumentModel.id.in_(clar_doc_ids))
                )
                for doc_id, filename in res_docs:
                    clar_doc_map[doc_id] = filename
        for cl in clarifications:
            doc_name = ""
            page_range = ""
            if cl.chunk_id:
                ch = clar_chunk_map.get(cl.chunk_id)
                if ch:
                    doc_name = clar_doc_map.get(ch.document_id, "")
                    if ch.page_start is not None and ch.page_end is not None:
                        page_range = f"{ch.page_start}-{ch.page_end}"
                    elif ch.page_start is not None:
                        page_range = str(ch.page_start)
            ws_clar.append(
                [
                    str(cl.id),
                    doc_name,
                    page_range,
                    cl.question_text,
                    cl.answer_text or "",
                    cl.status,
                    cl.created_at.isoformat(),
                    cl.answered_at.isoformat() if cl.answered_at else "",
                ]
            )
    # Auto-adjust widths for clarification sheet
    for i, column_title in enumerate(header2, start=1):
        max_length = len(column_title)
        for cell in ws_clar[get_column_letter(i)]:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws_clar.column_dimensions[get_column_letter(i)].width = max_length + 2

    # Serialize workbook to bytes
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()